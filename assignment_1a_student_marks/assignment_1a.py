import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Input and Output
INPUT_XLSX = Path("student.xlsx")     # input file (raw data)
OUTPUT_XLSX = Path("students.xlsx")  # final workbook file

# ----------------- COMPUTE SUMMARY -----------------
def compute_summary(df: pd.DataFrame) -> pd.DataFrame:
    subjects = ["Math", "Physics", "Chemistry", "Biology"]
    df_subjects = df[subjects].apply(pd.to_numeric, errors="coerce").fillna(0.0)

    total = df_subjects.sum(axis=1).to_numpy()
    avg = (total / df_subjects.shape[1]).astype(float)

    # Grades using vectorized logic
    conditions = [
        avg >= 90,
        (avg >= 75) & (avg < 90),
        (avg >= 60) & (avg < 75),
        avg < 60,
    ]
    choices = ["A", "B", "C", "F"]
    grade = np.select(conditions, choices, default="F")

    summary = pd.DataFrame({
        "StudentID": df["StudentID"],
        "Name": df["Name"],
        "Total": total,
        "Average": avg.round(2),
        "Grade": grade,
    })
    return summary

# ----------------- TOP PERFORMERS -----------------
def top_performers(df: pd.DataFrame, k: int = 3) -> pd.DataFrame:
    subjects = ["Math", "Physics", "Chemistry", "Biology"]
    rows = []
    for subj in subjects:
        topk = df.nlargest(k, subj)[["StudentID", "Name", subj]].copy()
        topk.insert(0, "Subject", subj)
        rows.append(topk)
    return pd.concat(rows, ignore_index=True)

# ----------------- WRITE RESULTS -----------------
def write_results(df: pd.DataFrame, summary: pd.DataFrame, avg_by_subject: pd.Series, toppers: pd.DataFrame):
    # Create students.xlsx with all sheets
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="student")   # raw input data
        summary.to_excel(writer, index=False, sheet_name="Summary")
        toppers.to_excel(writer, index=False, sheet_name="Top Performers")
        avg_df = avg_by_subject.reset_index()
        avg_df.columns = ["Subject", "Average"]
        avg_df.to_excel(writer, index=False, sheet_name="Averages")

    # Add chart to Averages sheet
    wb = load_workbook(OUTPUT_XLSX)
    ws = wb["Averages"]

    chart = BarChart()
    chart.title = "Average Marks per Subject"
    chart.y_axis.title = "Average"
    chart.x_axis.title = "Subject"

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=2)  # 'Average' values
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)             # 'Subject' labels
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E2")

    wb.save(OUTPUT_XLSX)

# ----------------- MAIN -----------------
def main():
    if not INPUT_XLSX.exists():
        raise FileNotFoundError("Input Excel 'student.xlsx' not found. Provide your own input file.")

    df = pd.read_excel(INPUT_XLSX)

    # Compute
    summary = compute_summary(df)
    subjects = ["Math", "Physics", "Chemistry", "Biology"]
    avg_by_subject = df[subjects].apply(pd.to_numeric, errors="coerce").fillna(0.0).mean(axis=0)
    toppers = top_performers(df)

    # Persist
    write_results(df, summary, avg_by_subject, toppers)
    print("students.xlsx created with sheets: student, Summary, Top Performers, Averages (with chart).")

if __name__ == "__main__":
    main()
